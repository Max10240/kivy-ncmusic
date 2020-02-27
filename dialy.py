import logging
from time import time,sleep
import os, re, json,pickle

# File Options
class File():
    
    @staticmethod
    def save_to_file(obj, filename):
        method='wb' if type(obj)==bytes else 'w'
        with open(filename, method)as f:
            f.write(obj)

    @staticmethod
    def load_from_file(filename):
        try:
            with open(filename, 'r') as f:
                content=f.read()
        except UnicodeDecodeError:
            with open(filename, 'rb') as f:
                content=f.read()
        return content

    @staticmethod
    def save_to_json(obj, filename):
        with open(filename,'w')as f:
            json.dump(obj,f)

    @staticmethod
    def save_to_pickle(obj, filename):
        with open(filename, 'wb')as f:
            pickle.dump(obj, f)

    @staticmethod
    def load_from_json(filename):
        with open(filename, 'r')as f:
            return json.load(f)

    @staticmethod
    def load_from_pickle(filename):
        with open(filename, 'rb')as f:
            return pickle.load(f)

    @staticmethod
    def save_to_excel(data, fn='data.xlsx', first_col_name='序号', auto_adjust_col_width=True):
        import pandas
        
        # return if data is empty
        if not data:return
        
        result=pandas.DataFrame(columns=([first_col_name]+list(data[0].keys()))) if not os.path.exists(fn) else pandas.read_excel(fn)
        for i,d in enumerate(data, start=len(result[result.columns[0]])):
            dict_d={first_col_name:i}
            if d:
                dict_d.update({k:[v] for k,v in d.items()})
            else:
                continue
            result=result.append(pandas.DataFrame(dict_d, index=[0], ), )

        result=result.ix[:, [first_col_name]+list(data[0].keys())]
        result.to_excel(fn, index=False)
        
        if auto_adjust_col_width:File.adjust_col_width(fn)

    @staticmethod
    def adjust_col_width(fn):
        from openpyxl import load_workbook
        from openpyxl.utils import get_column_letter
        
        wb=load_workbook(fn)
        for sheet_name in wb.sheetnames:
            sheet=wb[sheet_name]
            col_max_width=[max([len(str(r.value))+sum(len(x) for x in re.findall('[^\x00-\xff]+', str(r.value))) if r.value else 0 for r in col]) for col in sheet.columns]
            
            for i,w in enumerate(col_max_width):
                col_letter=get_column_letter(i+1)
                sheet.column_dimensions[col_letter].width=max(min(w*2,100), 10)
                
        wb.save(filename=fn.replace('.', '_adj.'))

# Encryption Options
class Encrypt():
    import hashlib, base64
    
    @staticmethod
    def b64_encode(str_in):
        return base64.b64encode(str_in.encode('utf-8')).decode('utf-8')

    @staticmethod
    def b64_decode(str_in):
        return base64.b64decode(str_in).decode('utf-8')

    @staticmethod
    def md5(cnt):
        import hashlib
        hl = hashlib.md5()
        hl.update(str(cnt).encode('utf-8'))
        return hl.hexdigest()

# Time Options
def time_counter(n):
    now=time()
    def creat():
        nonlocal now
        if time()-now > n:
            now=time()
            return True
        return False
    return creat

class timer():
    n=0
    
    def __enter__(self):
        self.time_begin=time()
        
    def __exit__(self, type, value, traceback):
        timer.n=time()-self.time_begin
        
# List,str,dict Options
def unique_list(l, delete=1):
    u_l=[]
    for x in l:
        if x not in u_l:u_l.append(x)
    if delete:del(l)
    return u_l

def string_similar_rate(s1, s2):
    import difflib
    return difflib.SequenceMatcher(None, s1, s2).quick_ratio()
    


class IgnoreFilter(logging.Filter):
    def __init__(self, keyword):
        logging.Filter.__init__(self)
        self.keyword=keyword
    def filter(self,record):
        if record.stack_info:
            record.stack_info=re.sub('File.+idlelib/run.py.+?\n.+?\n','',
                                     record.stack_info, flags=re.S)
            record.stack_info=re.sub(' +(?=File)',' '*2,record.stack_info)
            record.stack_info=re.sub(' {2}File.+?\n.+?self\.logger\.warning\(msg,stack_info=stack_info.+?$',
                                     '',record.stack_info)
        return True
    
class Log():
    def __init__(self, name, level=logging.DEBUG):
        logger=logging.getLogger(name)
        logger.propagate=False
        logger.setLevel(level)
        ch = logging.StreamHandler()
        ch.setLevel(logging.DEBUG)
        formatter = logging.Formatter('>>> %(levelname)s:%(name)s:%(message)s')
        ch.setFormatter(formatter)
        ch.addFilter(IgnoreFilter(''))
        logger.addHandler(ch)
        self.logger=logger

    def debug(self,*args,**kw):
        msg='<->'.join([str(x) for x in args])
        self.logger.debug(msg,**kw)
        
    def info(self,*args,**kw, ):
        msg='<->'.join([str(x) for x in args])
        self.logger.info(msg,**kw)
        
    def warning(self,*args,stack_info=True,**kw,):
        msg='<->'.join([str(x) for x in args])
        self.logger.warning(msg,stack_info=stack_info,**kw)
        
    def error(self,*args,stack_info=True,**kw,):
        msg='<->'.join([str(x) for x in args])
        self.logger.error(msg,stack_info=stack_info,**kw)
        


